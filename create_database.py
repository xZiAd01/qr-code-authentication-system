import os
import random
import qrcode
from faker import Faker
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from datetime import datetime, timedelta

# Initialize Faker for generating fake names and data
fake = Faker()

# Define vehicle ownership distribution
vehicle_count = [1, 2, 3, 4]
weights = [0.65, 0.25, 0.08, 0.02]  # 65% have 1 car, 25% have 2, etc.

# Define possible vehicle statuses
vehicle_statuses = ["Operational", "Under Maintenance", "Awaiting Parts", "Idle"]
status_weights = [0.85, 0.10, 0.03, 0.02] # Most vehicles are operational

# Define common maintenance requirements
maintenance_reqs = [
    "Oil Change", "Tire Rotation", "Brake Inspection",
    "Fluid Check", "Battery Check", "Spark Plug Replacement",
    "No immediate requirements"
]
maintenance_req_weights = [0.2, 0.2, 0.15, 0.15, 0.1, 0.05, 0.15]

# Define possible car makes and models (updated as per request)
car_makes = ["Toyota", "Lexus"]
car_models = {
    "Toyota": [
        "Camry", "Corolla", "Yaris", "Avalon", "Supra",
        "RAV4", "Highlander", "Fortuner", "Land Cruiser", "Prado", "C-HR", "Rush",
        "Tacoma", "Tundra", "Hilux",
        "Innova", "Sienna", "Granvia",
        "Crown", "Veloz", "Urban Cruiser"
    ],
    "Lexus": [
        "ES", "IS", "LS", "RC", "LC",
        "RX", "NX", "UX", "GX", "LX",
        "LM"
    ]
}

print("Removing previous QR code images...")

# Define the folder to store QR code images
qr_folder = "QRcodes"

# If the folder doesn't exist, create it. Otherwise, delete all previous QR images inside
if not os.path.exists(qr_folder):
    os.makedirs(qr_folder)
else:
    for filename in os.listdir(qr_folder):
        file_path = os.path.join(qr_folder, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)

# This list will store all generated data for export
all_data = []

print("Generating fake customers, vehicles and QR codes...")

# Generate data for 500 fake customers
for _ in range(500):
    # Generate basic customer data
    first_name = fake.first_name()
    last_name = fake.last_name()
    customer_id = str(fake.unique.random_int(100000, 999999))
    expiration_date = str(fake.date_between(start_date='-5y', end_date='+5y'))

    # Generate and save a QR code image for the customer ID
    qr_path = f"{qr_folder}/{customer_id}.png"
    img = qrcode.make(customer_id)
    img.save(qr_path)

    # Randomly assign number of cars the customer owns
    car_num = random.choices(vehicle_count, weights=weights)[0]

    # Create a record for each car
    for _ in range(car_num):

        mileage = random.randint(10_000, 200_000)
        vin = fake.unique.vin()
        
        # Generate car make and model
        car_make = random.choice(car_makes)
        car_model = random.choice(car_models[car_make]) # Now uses car_make as key

        # Generate vehicle status and maintenance data
        status = random.choices(vehicle_statuses, weights=status_weights)[0]
        
        # Maintenance Dates
        today = datetime.today()
        last_maintenance_date = today - timedelta(days=random.randint(30, 365))  # 1 month to 1 year ago
        next_maintenance_due = today + timedelta(days=random.randint(30, 180))  # 1 month to 6 months ahead

        # Convert to 'Y-m-d' format
        last_maintenance_date = last_maintenance_date.strftime("%Y-%m-%d")
        next_maintenance_due = next_maintenance_due.strftime("%Y-%m-%d")
        
        maintenance_requirements = random.choices(maintenance_reqs, weights=maintenance_req_weights)[0]

        # Simulate predictive maintenance alert and remaining useful life
        predictive_maintenance_alert = random.choice([True, False, False, False, False]) # Mostly False
        if predictive_maintenance_alert:
            # Remaining Life (km) based on mileage (more mileage = less remaining life)
            if mileage < 50_000:
                remaining_life = random.randint(8000, 15000)
            elif mileage < 100_000:
                remaining_life = random.randint(5000, 12000)
            else:
                remaining_life = random.randint(1000, 8000)
        else:
            remaining_life = "N/A"

        # Append the record as a dictionary
        all_data.append({
            'customer_id': customer_id,
            'first_name': first_name,
            'last_name': last_name,
            'number of cars': car_num,
            'car_vin': vin,
            'car_make': car_make, # Changed from car_type
            'car_model': car_model,
            'expiration_date': expiration_date,
            'status': status,
            'last_maintenance_date': str(last_maintenance_date),
            'next_maintenance_due': str(next_maintenance_due),
            'maintenance_requirements': maintenance_requirements,
            'predictive_maintenance_alert': predictive_maintenance_alert,
            'remaining_useful_life': remaining_life,
            'qr_path': qr_path  # Path to the QR code image file
        })

print("Adding data to Excel file...")

# Create a new Excel workbook and select the default worksheet
wb = Workbook()
ws = wb.active

# Define and write the column headers, including new fields
headers = [
    'customer_id', 'first_name', 'last_name', 'number of cars', 
    'car_vin', 'car_make', 'car_model',
    'expiration_date', 'status', 'last_maintenance_date', 'next_maintenance_due',
    'maintenance_requirements', 'predictive_maintenance_alert', 'remaining_useful_life',
    'QR_code'
]
ws.append(headers)

# Center alignment for all cell values
center_alignment = Alignment(horizontal='center', vertical='center')

# Loop over each row of data to populate the Excel sheet
for row in all_data:
    # Insert values except the image
    ws.append([
        row['customer_id'],
        row['first_name'],
        row['last_name'],
        row['number of cars'],
        row['car_vin'],
        row['car_make'], # Changed from car_type
        row['car_model'],
        row['expiration_date'],
        row['status'],
        row['last_maintenance_date'],
        row['next_maintenance_due'],
        row['maintenance_requirements'],
        row['predictive_maintenance_alert'],
        row['remaining_useful_life'],
        '',  # Placeholder for image
    ])

    # Adjust row height to fit QR image
    ws.row_dimensions[ws.max_row].height = 80

    # Create an image object and resize it
    img = Image(row['qr_path'])
    img.width = img.height = 60

    # Calculate target cell (last column, current row)
    img_cell = f'{chr(ord("A") + len(headers) - 1)}{ws.max_row}' # Dynamically get last column letter

    # Adjust column width and row height to make image appear centered
    ws.column_dimensions[chr(ord("A") + len(headers) - 1)].width = img.width * 0.14  # ~11.2 width for 80px
    ws.row_dimensions[ws.max_row].height = img.height * 0.75  # ~60 height for 80px

    # Insert the image into the worksheet
    ws.add_image(img, img_cell)

print("Resizing columns...")

# Auto-adjust all columns and apply alignment
for col_idx, col in enumerate(ws.columns):
    # Skip QR_code column for auto-adjusting width as it's set specifically for image
    if col_idx == len(headers) - 1:
        continue
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    adjusted_width = max_length + 2  # Extra space for readability
    ws.column_dimensions[col[0].column_letter].width = adjusted_width
    for cell in col:
        cell.alignment = center_alignment  # Apply center alignment

# Save the workbook as an Excel file
wb.save("customers_with_vehicles.xlsx")

print("DONE, All data is successfully generated and stored in (customers_with_vehicles.xlsx)")